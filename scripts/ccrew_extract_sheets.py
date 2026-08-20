"""Dumb extraction of the three historical CCrew Google Sheets to JSON.

Deliberately applies NO rules — every threshold/eligibility/staff decision is
made by lib/ccrew/rules.js so the backfill and the live upload path can never
drift apart. This only knows where the cells are.

Sheets are link-accessible with no auth:
  https://docs.google.com/spreadsheets/d/<id>/export?format=xlsx
"""
import json, re, sys, urllib.request, io
import openpyxl

DOCS = {
    "2024": "1Kir7Mmtyrn4TYMYayBau_GSmA76hQFGlfBPF_xEYdkE",
    "2025": "1b2FXO_W65a8HpHJUvoOvfT4Zt2JJ7BDbArc3q2p2Zy4",
    "2026": "1fzdS4pnatgs4Cqp914NBahH9U3bV6EZ7H1Yf-qlFqwo",
}

# The three 2024 tabs use DIFFERENT column orders. 1-indexed offsets from
# column B (column A is a row number, or blank). Ignore every other tab in
# that doc (Template, Master, Nov #1/#2, Dec #1/#2) — an unrelated biweekly
# experiment.
TABS_2024 = {
    "Oct Committed Crew":  {"period": "2024-10-01", "att": 2, "exp": 3, "pkg": 5},
    "Nov Committed Crew":  {"period": "2024-11-01", "att": 2, "exp": 5, "pkg": 3},
    " Dec Committed Crew": {"period": "2024-12-01", "att": 2, "exp": 5, "pkg": 3},
}

MONTHS = ["January","February","March","April","May","June",
          "July","August","September","October","November","December"]


def load(doc_id):
    with urllib.request.urlopen(f"https://docs.google.com/spreadsheets/d/{doc_id}/export?format=xlsx") as r:
        return openpyxl.load_workbook(io.BytesIO(r.read()), data_only=True)


def clean(v):
    return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""


def num(v):
    return int(v) if isinstance(v, (int, float)) else None


def main():
    out = {"months": {}}

    wb = load(DOCS["2024"])
    for tab, L in TABS_2024.items():
        rows = []
        for r in wb[tab].iter_rows(min_row=2, values_only=True):
            if not r[1]:
                continue
            rows.append({
                "name": clean(r[1]),
                "attendance": num(r[L["att"]]) or 0,
                "expected": num(r[L["exp"]]),
                "packages": clean(r[L["pkg"]]),
            })
        # RAW roster: roughly half of each of these tabs is below 80%. The
        # rule has to be applied — taking them at face value inflates streaks.
        out["months"][L["period"]] = {"kind": "roster", "tab": tab, "rows": rows}

    for year in ("2025", "2026"):
        wb = load(DOCS[year])
        for i, m in enumerate(MONTHS):
            if m not in wb.sheetnames:
                continue
            rows = []
            for r in wb[m].iter_rows(values_only=True):
                if not r[0]:
                    continue
                rows.append({
                    "name": clean(r[0]),
                    "attendance": num(r[1]) or 0,
                    "packages": clean(r[2]),
                    "expected": num(r[3]),
                })
            if not rows:
                continue
            # Already FILTERED: these tabs are the finished crew lists, so
            # membership of the tab IS qualification. The ratio column holds
            # hardcoded values with known typos and must not be re-filtered on.
            out["months"][f"{year}-{i+1:02d}-01"] = {"kind": "crew", "tab": f"{year} {m}", "rows": rows}

    out["months"] = dict(sorted(out["months"].items()))
    json.dump(out, sys.stdout, indent=1)


if __name__ == "__main__":
    main()
